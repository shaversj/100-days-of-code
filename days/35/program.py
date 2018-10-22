import openpyxl

# load excel spreadsheet
wb = openpyxl.load_workbook('duesRecords.xlsx')

# Get sheet
sheet = wb.active

# Determines the number associated with the last column
lastCol = sheet.max_column

# Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email
